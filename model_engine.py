"""
model_engine.py — Regional Economic Impact Model v7.2
======================================================
Pure-computation module. No Streamlit imports. No side effects.

Methodology
-----------
Leontief I-O framework, Type I and Type II multipliers.
  - 17 model sectors aggregated from BEA 402-industry Benchmark tables.
  - Supply-Demand Pool (SDP) regionalization as primary method
    (Kronenberg 2009; Flegg & Tohmo 2016). SLQ and FLQ retained for
    sensitivity comparison only.
  - Employment coefficients in jobs per $1M value added (BEA Table 1).
  - Household closure for induced effects (Type II).
  - ±30% uncertainty bands on all outputs (Watson et al. 2015).

Data sources
------------
  - BEA 2017 Benchmark I-O tables (uploaded at runtime): Z, D, B_domestic
  - BLS QCEW county file (uploaded at runtime): employment, wages, LQs
  - BEA GDP-by-Industry 2022 (CSV primary, API secondary): VA, GO
  - BLS CES 2022 (live API, fallback to published constants): national emp
  - BEA NIPA Table T20600 (live API, fallback to user slider): MPC

v7.2 changes from v7.1
-----------------------
  - Type hints added throughout.
  - Docstrings improved: units, sources, and edge cases documented.
  - Input validation added to major public functions.
  - Diagnostic metadata expanded in build_model return dict.
  - Numerical safeguards: near-singular inversion guard, explicit
    spectral-radius pre-check before inversion.
  - load_qcew: required-column check and graceful wage imputation.
  - _load_csv_cache: explicit missing-column warning.
  - BLS fallback: logs which series used live vs constant.
  - COMP_VA_RATIOS promoted to module-level constant (was recomputed
    on every build_model call).
  - Minor variable naming improvements for clarity.
  - No changes to methodology, sector structure, or public API.
"""

from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import openpyxl
import requests

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

VERSION: str = "7.3"

_HERE: Path = Path(__file__).parent
DATA_DIR: Path = _HERE / "data"

# ---------------------------------------------------------------------------
# Sector definitions — 17 sectors, indices 0–16
# ---------------------------------------------------------------------------

SECTOR_LABELS: list[str] = [
    "Agriculture",            # 0  NAICS 11
    "Mining & Utilities",     # 1  NAICS 21, 22
    "Construction",           # 2  NAICS 23
    "Mfg — Durable",          # 3  NAICS 33DG
    "Mfg — Non-Durable",      # 4  NAICS 31ND
    "Wholesale Trade",        # 5  NAICS 42
    "Retail Trade",           # 6  NAICS 44-45
    "Transportation & Whsg.", # 7  NAICS 48-49
    "Information",            # 8  NAICS 51
    "Finance & Real Estate",  # 9  NAICS 52-53
    "Professional Services",  # 10 NAICS 54-55
    "Admin & Waste Services", # 11 NAICS 56
    "Education & Health",     # 12 NAICS 61-62
    "Arts & Accommodation",   # 13 NAICS 71-72
    "Private Other Services", # 14 NAICS 81          (own_code 5)
    "Federal Gov & Defense",  # 15 NAICS 921         (own_code 1)
    "State & Local Gov",      # 16 NAICS 922-928     (own_code 2+3)
]
N: int = len(SECTOR_LABELS)  # 17

# Chart colours — one per sector, stable across runs
SECTOR_COLORS: list[str] = [
    "#4ade80", "#a78bfa", "#f97316", "#60a5fa", "#f472b6",
    "#34d399", "#fb923c", "#818cf8", "#2dd4bf", "#c084fc",
    "#e879f9", "#38bdf8", "#facc15", "#f87171", "#94a3b8",
    "#1d4ed8", "#15803d",
]

# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------

BEA_BASE: str = "https://apps.bea.gov/api/data"
BLS_BASE: str = "https://api.bls.gov/publicAPI/v2/timeseries/data/"

# ---------------------------------------------------------------------------
# Multiplier plausibility bounds
# Source: BEA RIMS II User Handbook (2022); IMPLAN validation studies
# ---------------------------------------------------------------------------

MULT_TYPE1_MIN: float = 1.1
MULT_TYPE1_MAX: float = 2.5
MULT_TYPE2_MIN: float = 1.2
MULT_TYPE2_MAX: float = 3.0

# ---------------------------------------------------------------------------
# Uncertainty
# Source: Watson et al. (2015) J. of Regional Science 55(1)
#   ±20–30% gap between I-O predictions and observed outcomes
#   + ~±10–15% for applying 2017 production structure to current year
# ---------------------------------------------------------------------------

UNCERTAINTY_PCT: float = 0.30

# ---------------------------------------------------------------------------
# Compensation / Value-Added ratios by sector
# Computed once from the 2022 BEA hardcoded fallback arrays.
# Used to derive compensation when only VA is available.
# These are structural parameters (stable across years); not year-specific.
# ---------------------------------------------------------------------------

# Hardcoded 2022 fallback arrays ($M, 17 elements)
# Source: BEA GDP-by-Industry 2022, aggregated via BEA_SUMMARY_TO_SECTOR crosswalk.
# These are used only when the CSV cache is absent AND the live API fails.
# Units: millions of US dollars ($M)

BEA_OUTPUT_FALLBACK: np.ndarray = np.array([
    464_200, 1_031_800, 1_872_400, 3_421_600, 3_190_200,
    2_391_700, 2_088_400, 1_512_300, 2_155_700, 7_321_500,
    3_289_100, 1_256_100, 3_298_400, 1_702_800,
    812_000, 1_083_000, 3_263_300,
], dtype=float)

BEA_COMP_FALLBACK: np.ndarray = np.array([
    58_200, 174_700, 617_400, 841_800, 402_700,
    618_800, 686_900, 473_100, 426_800, 1_063_500,
    1_562_300, 438_100, 1_352_300, 578_800,
    260_000, 480_000, 1_832_600,
], dtype=float)

BEA_VA_FALLBACK: np.ndarray = np.array([
    233_700, 492_000, 993_400, 1_371_800, 790_200,
    1_202_800, 1_213_700, 753_100, 1_198_400, 5_068_700,
    2_204_500, 731_800, 2_244_400, 994_200,
    600_000, 900_000, 2_362_400,
], dtype=float)

# Sector-level comp/VA ratios derived from the fallback arrays above.
# Promoted to module level so they are computed once, not on every build call.
_COMP_VA_RATIOS: np.ndarray = np.where(
    BEA_VA_FALLBACK > 0,
    BEA_COMP_FALLBACK / BEA_VA_FALLBACK,
    0.25,  # default ratio for any sector with zero VA in the fallback
)

# ---------------------------------------------------------------------------
# Sector crosswalk — BEA industry codes → 17-sector indices
# ---------------------------------------------------------------------------

# BEA special non-numeric commodity codes
_SPECIAL: dict[str, int] = {
    # Scrap, used goods, adjustments → Private Other Services (14)
    "S00500": 14, "S00600": 14, "S00101": 14, "S00102": 14,
    "S00201": 14, "S00202": 14, "S00203": 14,
    # State/local government → State & Local Gov (16)
    "GSLGE": 16, "GSLGH": 16, "GSLGO": 16,
    "GSLE": 16, "GSLO": 16,
    # Federal government → Federal Gov & Defense (15)
    "GFGD": 15, "GFGN": 15, "GFE": 15,
    # Non-comparable imports / rest-of-world → exclude (-1)
    "S00401": -1, "S00402": -1, "S00300": -1, "S00900": -1,
    # Real estate sub-codes → Finance & Real Estate (9)
    "531HSO": 9, "531HSR": 9, "532RL": 9, "ORE000": 9,
    "HS0000": 9, "521CI0": 9, "525000": 9,
    # Retail / transport sub-codes
    "4B0000": 6, "487OS0": 7, "488A00": 7,
    # Agriculture sub-codes
    "111CA0": 0, "113FF0": 0,
}

# Non-durable manufacturing 3-digit NAICS codes
_NONDUR: frozenset[str] = frozenset({
    "311", "312", "313", "314", "315", "316",
    "322", "323", "324", "325", "326",
})

# NAICS code → sector index mapping (2- and 3-digit keys)
NAICS_TO_SECTOR: dict[str, int] = {
    # Agriculture
    "11": 0, "111": 0, "112": 0, "113": 0, "114": 0, "115": 0,
    # Mining & Utilities
    "21": 1, "211": 1, "212": 1, "213": 1, "22": 1, "221": 1,
    # Construction
    "23": 2, "236": 2, "237": 2, "238": 2,
    # Manufacturing (durable default; non-durables override below via _NONDUR)
    "31": 3, "32": 4, "33": 3,
    "311": 4, "312": 4, "313": 4, "314": 4, "315": 4, "316": 4,
    "321": 3, "322": 4, "323": 4, "324": 4, "325": 4, "326": 4,
    "327": 3, "331": 3, "332": 3, "333": 3, "334": 3, "335": 3,
    "336": 3, "337": 3, "339": 3,
    # Wholesale
    "42": 5, "423": 5, "424": 5, "425": 5,
    # Retail
    "44": 6, "45": 6,
    # Transportation & Warehousing
    "48": 7, "49": 7,
    # Information
    "51": 8,
    # Finance & Real Estate
    "52": 9, "53": 9,
    # Professional Services
    "54": 10, "55": 10,
    # Admin & Waste Services
    "56": 11,
    # Education & Health
    "61": 12, "62": 12,
    # Arts & Accommodation
    "71": 13, "72": 13,
    # Private Other Services
    "81": 14, "811": 14, "812": 14, "813": 14, "814": 14,
    # Government — split by sub-NAICS
    "921": 15,
    "922": 16, "923": 16, "924": 16, "925": 16,
    "926": 16, "927": 16, "928": 16,
    "92": 16,  # generic government defaults to state/local
}

# BEA summary industry codes → sector index
# Used when parsing GDP-by-Industry API responses
BEA_SUMMARY_TO_SECTOR: dict[str, int] = {
    "111CA": 0, "113FF": 0,
    "211": 1, "212": 1, "213": 1, "22": 1,
    "23": 2,
    "321": 3, "327": 3, "331": 3, "332": 3, "333": 3, "334": 3,
    "335": 3, "3361MV": 3, "3364OT": 3, "337": 3, "339": 3,
    "311FT": 4, "313TT": 4, "315AL": 4, "322": 4, "323": 4,
    "324": 4, "325": 4, "326": 4,
    "42": 5,
    "441": 6, "445": 6, "452": 6, "4A0": 6,
    "481": 7, "482": 7, "483": 7, "484": 7, "485": 7,
    "486": 7, "487OS": 7, "493": 7,
    "511": 8, "512": 8, "513": 8, "514": 8,
    "521CI": 9, "523": 9, "524": 9, "525": 9,
    "HS": 9, "ORE": 9, "532RL": 9,
    "5411": 10, "5412OP": 10, "5415": 10, "55": 10,
    "561": 11, "562": 11,
    "61": 12, "621": 12, "622": 12, "623": 12, "624": 12,
    "711AS": 13, "713": 13, "721": 13, "722": 13,
    "81": 14,
    "GFGD": 15, "GFGN": 15, "GFE": 15,
    "GSLE": 16, "GSLO": 16,
}

# ---------------------------------------------------------------------------
# Crosswalk helpers
# ---------------------------------------------------------------------------

def naics_to_sector(code: str) -> int:
    """
    Map a 2- to 4-digit NAICS code string to a model sector index (0–16).

    Tries the first 3 digits, then the first 2. Raises ValueError if
    the code is not recognised. Callers should catch ValueError for
    user-facing NAICS entry.

    Parameters
    ----------
    code : str
        NAICS code, e.g. "23", "336", "9211".

    Returns
    -------
    int
        Sector index in [0, N-1].

    Raises
    ------
    ValueError
        If the code cannot be mapped to any sector.
    """
    code = str(code).strip()
    if not code:
        raise ValueError("NAICS code must not be empty.")
    for length in (3, 2):
        k = code[:length]
        if k in NAICS_TO_SECTOR:
            return NAICS_TO_SECTOR[k]
    raise ValueError(
        f"NAICS code '{code}' not recognised. "
        "Check that the code is a valid 2–4 digit NAICS value."
    )


def _code_to_sector(code: str) -> int:
    """
    Map a raw BEA industry code to a sector index.

    Returns -1 for explicitly excluded codes (rest-of-world, non-comparable
    imports). Returns 14 (Private Other Services) for unrecognised codes
    as a conservative default.

    Parameters
    ----------
    code : str
        BEA commodity or industry code, e.g. "23", "GFGD", "531HSO".
    """
    code = str(code).strip()
    if code in _SPECIAL:
        r = _SPECIAL[code]
        return r if r is not None else -1
    digits = "".join(c for c in code if c.isdigit())
    if not digits:
        return 14  # non-numeric, unrecognised → private other services
    p2, p3 = digits[:2], digits[:3]
    if   p2 == "11":              return 0
    elif p2 in ("21", "22"):      return 1
    elif p2 == "23":              return 2
    elif p2 in ("31", "32", "33"):return 4 if p3 in _NONDUR else 3
    elif p2 == "42":              return 5
    elif p2 in ("44", "45"):      return 6
    elif p2 in ("48", "49"):      return 7
    elif p2 == "51":              return 8
    elif p2 in ("52", "53"):      return 9
    elif p2 in ("54", "55"):      return 10
    elif p2 == "56":              return 11
    elif p2 in ("61", "62"):      return 12
    elif p2 in ("71", "72"):      return 13
    elif p2 == "81":              return 14
    elif p2 == "92":
        return 15 if p3 == "921" else 16
    return 14  # unrecognised numeric prefix → private other services


# ---------------------------------------------------------------------------
# BEA file parsers
# ---------------------------------------------------------------------------

def _safe_float(v: Any) -> float:
    """Return float(v) or 0.0 for None / non-numeric cell values."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def parse_use_table(src: Any, sheet: str = "2017") -> dict:
    """
    Parse the BEA 402-industry Use Table workbook.

    Expects the 'After Redefinitions (Producer Value)' file:
    IOUse_After_Redefinitions_PRO_Detail.xlsx

    Layout assumptions (verified against BEA 2017 Benchmark release):
      - Column indices 2–403 (0-based): 402 industry columns
      - Row indices 6–407: 402 commodity rows (Z matrix)
      - Row 409: VA compensation; 410: VA taxes; 411: VA GOS; 412: VA total
      - Row 413: total industry output (x vector)
      - Column 405: Personal Consumption Expenditure (PCE)
      - Row 4: industry names header; Row 5: industry codes header

    Parameters
    ----------
    src : file-like or path
        Opened file object or path accepted by openpyxl.load_workbook.
    sheet : str
        Worksheet name. Default '2017' for BEA 2017 Benchmark.

    Returns
    -------
    dict with keys:
      Z          : np.ndarray (402, 402)  — intermediate use matrix ($M)
      x          : np.ndarray (402,)      — total output by industry ($M)
      va_comp    : np.ndarray (402,)      — compensation of employees ($M)
      va_taxes   : np.ndarray (402,)      — taxes on production ($M)
      va_gos     : np.ndarray (402,)      — gross operating surplus ($M)
      va_total   : np.ndarray (402,)      — total value added ($M)
      pce        : np.ndarray (402,)      — PCE final demand ($M)
      ind_codes  : list[str] (402)        — BEA industry codes
      ind_names  : list[str] (402)        — BEA industry names
      com_codes  : list[str] (402)        — BEA commodity codes
    """
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found in Use Table workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    C0, C1 = 2, 404   # industry column range (0-based, exclusive end)
    R0, R1 = 6, 408   # commodity row range
    PCE_COL = 405

    def _row_to_array(row_idx: int) -> np.ndarray:
        return np.array(
            [_safe_float(rows[row_idx][c]) for c in range(C0, C1)],
            dtype=float,
        )

    Z = np.zeros((R1 - R0, C1 - C0), dtype=float)
    for ri, r in enumerate(range(R0, R1)):
        for ci, c in enumerate(range(C0, C1)):
            Z[ri, ci] = _safe_float(rows[r][c])

    pce = np.array(
        [_safe_float(rows[r][PCE_COL]) for r in range(R0, R1)],
        dtype=float,
    )

    return {
        "Z":         Z,
        "x":         _row_to_array(413),
        "va_comp":   _row_to_array(409),
        "va_taxes":  _row_to_array(410),
        "va_gos":    _row_to_array(411),
        "va_total":  _row_to_array(412),
        "pce":       pce,
        "ind_codes": [str(rows[5][c]) for c in range(C0, C1)],
        "ind_names": [str(rows[4][c]) for c in range(C0, C1)],
        "com_codes": [str(rows[r][0]) for r in range(R0, R1)],
    }


def parse_D_matrix(src: Any, sheet: str = "2017") -> dict:
    """
    Parse the BEA Market Share Matrix D (IxC_MS_Detail.xlsx).

    D is a 402×402 commodity-by-industry market share matrix where each
    column sums to 1.0. Columns that deviate from 1.0 by more than 0.01
    are rescaled.

    Parameters
    ----------
    src : file-like or path
    sheet : str
        Worksheet name. Default '2017'.

    Returns
    -------
    dict with keys:
      D          : np.ndarray (402, 402) — market share matrix
      ind_codes  : list[str] (402)       — BEA industry codes (row labels)
    """
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found in Market Share Matrix workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ind_codes = [str(rows[r][0]) for r in range(6, 408)]
    D = np.zeros((402, 402), dtype=float)
    for ri, r in enumerate(range(6, 408)):
        for ci, c in enumerate(range(2, 404)):
            D[ri, ci] = _safe_float(rows[r][c])

    # Rescale columns that do not sum to 1.0 (handles minor rounding in BEA file)
    col_sums = D.sum(axis=0)
    bad = np.abs(col_sums - 1.0) > 0.01
    if bad.any():
        D[:, bad] /= np.where(col_sums[bad] > 0, col_sums[bad], 1.0)

    return {"D": D, "ind_codes": ind_codes}


def parse_B_domestic(src: Any, sheet: str = "2017") -> np.ndarray:
    """
    Parse the BEA Domestic Direct Requirements Matrix B
    (CxI_Domestic_DR_Detail.xlsx).

    B is a 402×402 commodity-by-industry direct requirements matrix.
    Each column represents the commodity inputs required per dollar of
    industry output.

    Parameters
    ----------
    src : file-like or path
    sheet : str
        Worksheet name. Default '2017'.

    Returns
    -------
    np.ndarray (402, 402) — domestic direct requirements matrix (dimensionless)
    """
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet}' not found in Domestic Requirements workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    B = np.zeros((402, 402), dtype=float)
    for ri, r in enumerate(range(5, 407)):
        for ci, c in enumerate(range(2, 404)):
            B[ri, ci] = _safe_float(rows[r][c])
    return B


# ---------------------------------------------------------------------------
# QCEW county file parser
# ---------------------------------------------------------------------------

_QCEW_REQUIRED_COLS: frozenset[str] = frozenset({
    "own_code", "agglvl_code", "industry_code",
    "annual_avg_emplvl", "lq_annual_avg_emplvl",
    "avg_annual_pay", "total_annual_wages",
    "area_fips", "year",
})

# NAICS 2-digit / hyphenated codes to sector index (private establishments)
_QCEW_NAICS2_MAP: dict[str, int] = {
    "11": 0, "22": 1, "23": 2, "31-33": 3,
    "42": 5, "44-45": 6, "48-49": 7,
    "51": 8, "52": 9, "53": 9, "54": 10, "55": 10, "56": 11,
    "61": 12, "62": 12, "71": 13, "72": 13,
    "81": 14, "99": 14,
}

_QCEW_NONDUR_3D: frozenset[str] = frozenset({
    "311", "312", "313", "314", "315", "316",
    "322", "323", "324", "325", "326",
})


# ---------------------------------------------------------------------------
# Proprietor employment ratios by sector
# Source: BEA NIPA Table 6.4 (2022) — proprietors as share of total employment
# Applied to QCEW wage-and-salary base to estimate self-employed workers
# not captured by unemployment insurance payroll records.
# Government sectors (15, 16) have zero proprietors by definition.
# ---------------------------------------------------------------------------

_PROPRIETOR_RATIOS: np.ndarray = np.array([
    0.642,  #  0 Agriculture           — farm proprietors very high
    0.052,  #  1 Mining & Utilities
    0.241,  #  2 Construction          — many independent contractors
    0.020,  #  3 Mfg — Durable
    0.021,  #  4 Mfg — Non-Durable
    0.062,  #  5 Wholesale Trade
    0.082,  #  6 Retail Trade
    0.138,  #  7 Transportation        — owner-operator truckers
    0.093,  #  8 Information
    0.211,  #  9 Finance & Real Estate — agents, advisors
    0.298,  # 10 Professional Services — consultants, lawyers, accountants
    0.087,  # 11 Admin & Waste Services
    0.111,  # 12 Education & Health
    0.091,  # 13 Arts & Accommodation
    0.187,  # 14 Private Other Services
    0.000,  # 15 Federal Gov & Defense  — no proprietors
    0.000,  # 16 State & Local Gov      — no proprietors
], dtype=float)


def load_qcew(src: Any) -> dict:
    """
    Parse a BLS QCEW county-level annual CSV file.

    Handles the 17-sector government split:
      own_code 1 → Federal Gov & Defense (sector 15)
      own_code 2 → State & Local Gov    (sector 16, state portion)
      own_code 3 → State & Local Gov    (sector 16, local portion)
      own_code 5, NAICS 81 → Private Other Services (sector 14)

    Manufacturing is split into durable (sector 3) and non-durable (sector 4)
    using 3-digit NAICS agglvl_code=75 rows where available.

    Proprietor employment is added on top of QCEW wage-and-salary employment
    using BEA national proprietor ratios (NIPA Table 6.4, 2022). This
    corrects the known QCEW undercount of self-employed workers who are
    excluded from unemployment insurance reporting.

    Parameters
    ----------
    src : file-like or path
        BLS QCEW county CSV, e.g. '51199.csv' for York County VA.

    Returns
    -------
    dict with keys:
      emp              : np.ndarray (17,)  — total employment including proprietors
      emp_ws           : np.ndarray (17,)  — wage-and-salary only (QCEW direct)
      emp_prop         : np.ndarray (17,)  — estimated proprietors added
      lq               : np.ndarray (17,)  — employment location quotient
      wage             : np.ndarray (17,)  — avg annual pay ($ per worker)
      year             : int               — data year
      fips             : str               — 5-digit FIPS code
      total_county_emp : float             — total employment including proprietors
    """
    df = pd.read_csv(src, dtype={"area_fips": str, "industry_code": str})
    df["industry_code"] = df["industry_code"].str.strip()

    # Validate required columns
    missing_cols = _QCEW_REQUIRED_COLS - set(df.columns)
    if missing_cols:
        raise ValueError(
            f"QCEW file is missing required columns: {sorted(missing_cols)}. "
            "Ensure you are using the BLS QCEW annual CSV format."
        )

    year = int(df["year"].max())

    emp_ws = np.zeros(N, dtype=float)  # wage-and-salary (QCEW direct)
    lq_n   = np.zeros(N, dtype=float)  # employment-weighted LQ numerator
    wage_n = np.zeros(N, dtype=float)  # employment-weighted wage numerator

    # Private establishments, 2-digit NAICS (agglvl_code 74)
    priv2 = df[(df["own_code"] == 5) & (df["agglvl_code"] == 74)].copy()
    for _, row in priv2.iterrows():
        s = _QCEW_NAICS2_MAP.get(row["industry_code"], -1)
        if s < 0:
            continue
        e = float(row["annual_avg_emplvl"])
        emp_ws[s]  += e
        lq_n[s]    += float(row["lq_annual_avg_emplvl"]) * e
        wage_n[s]  += float(row["avg_annual_pay"]) * e

    # Manufacturing durable / non-durable split using 3-digit NAICS rows
    priv3 = df[(df["own_code"] == 5) & (df["agglvl_code"] == 75)].copy()
    mfg3 = priv3[priv3["industry_code"].str.match(r"^3[123]$|^3[123]\d$")]
    if len(mfg3) > 0:
        dur    = mfg3[~mfg3["industry_code"].isin(_QCEW_NONDUR_3D)]
        nondur = mfg3[ mfg3["industry_code"].isin(_QCEW_NONDUR_3D)]
        for s_idx in (3, 4):
            emp_ws[s_idx] = lq_n[s_idx] = wage_n[s_idx] = 0.0
        for df_sub, s_idx in ((dur, 3), (nondur, 4)):
            e_tot = float(df_sub["annual_avg_emplvl"].sum())
            if e_tot > 0:
                emp_ws[s_idx]  = e_tot
                lq_n[s_idx]    = (
                    df_sub["lq_annual_avg_emplvl"]
                    * df_sub["annual_avg_emplvl"]
                ).sum()
                wage_n[s_idx]  = (
                    df_sub["avg_annual_pay"]
                    * df_sub["annual_avg_emplvl"]
                ).sum()

    # Government split — county-wide totals (agglvl_code 71)
    for own_code, sector_idx in ((1, 15), (2, 16), (3, 16)):
        gov_rows = df[(df["own_code"] == own_code) & (df["agglvl_code"] == 71)]
        if len(gov_rows) == 0:
            continue
        e_tot = float(gov_rows["annual_avg_emplvl"].sum())
        emp_ws[sector_idx]  += e_tot
        wage_n[sector_idx]  += float(gov_rows["total_annual_wages"].sum())
        lq_n[sector_idx]    += (
            gov_rows["lq_annual_avg_emplvl"] * gov_rows["annual_avg_emplvl"]
        ).sum()

    lq   = np.where(emp_ws > 0, lq_n   / emp_ws, 0.0)
    wage = np.where(emp_ws > 0, wage_n / emp_ws, 0.0)

    # ── Add proprietor employment ─────────────────────────────────────────────
    # Formula: proprietors = WS_emp × ratio / (1 - ratio)
    # where ratio = proprietors / total_employment (from BEA national averages)
    # Government sectors have ratio=0 so no proprietors are added there.
    with np.errstate(divide="ignore", invalid="ignore"):
        emp_prop = np.where(
            _PROPRIETOR_RATIOS > 0,
            np.round(emp_ws * _PROPRIETOR_RATIOS / (1.0 - _PROPRIETOR_RATIOS)),
            0.0,
        )
    emp_prop = np.maximum(emp_prop, 0.0)  # defensive floor
    emp_total = emp_ws + emp_prop

    return {
        "emp":              emp_total,   # total including proprietors
        "emp_ws":           emp_ws,      # wage-and-salary only (QCEW direct)
        "emp_prop":         emp_prop,    # proprietors added
        "lq":               lq,          # based on W&S (as published by BLS)
        "wage":             wage,
        "year":             year,
        "fips":             str(df["area_fips"].iloc[0]),
        "total_county_emp": float(emp_total.sum()),
    }


# ---------------------------------------------------------------------------
# Matrix builders
# ---------------------------------------------------------------------------

def build_A_domestic(
    D: np.ndarray,
    B_dom: np.ndarray,
    use_data: dict,
) -> dict:
    """
    Construct the 402×402 domestic technical coefficients matrix A.

    A = D @ B_domestic

    Columns that sum to ≥ 1.0 are rescaled to 0.99 to ensure Leontief
    stability. Value-added and labor-income shares are derived from the
    Use Table.

    Parameters
    ----------
    D       : (402, 402) market share matrix from parse_D_matrix
    B_dom   : (402, 402) domestic direct requirements from parse_B_domestic
    use_data : dict from parse_use_table

    Returns
    -------
    dict with keys:
      A        : np.ndarray (402, 402) — domestic technical coefficients
      va_share : np.ndarray (402,)     — value added / gross output
      li_share : np.ndarray (402,)     — compensation / gross output
    """
    A = D @ B_dom
    col_sums = A.sum(axis=0)
    unstable = col_sums >= 1.0
    if unstable.any():
        A[:, unstable] *= 0.99 / col_sums[unstable]

    x = use_data["x"]
    with np.errstate(divide="ignore", invalid="ignore"):
        va_share = np.where(x > 0, use_data["va_total"] / x, 0.0)
        li_share = np.where(x > 0, use_data["va_comp"]  / x, 0.0)

    return {"A": A, "va_share": va_share, "li_share": li_share}


def build_B_total(use_data: dict) -> np.ndarray:
    """
    Compute the total requirements matrix B_total = Z / x (column-wise).

    Used only for computing import-removed statistics; not used in the
    Leontief solve.

    Returns
    -------
    np.ndarray (402, 402) — total (domestic + imported) direct requirements
    """
    x = use_data["x"]
    with np.errstate(divide="ignore", invalid="ignore"):
        return np.where(x > 0, use_data["Z"] / x[np.newaxis, :], 0.0)


def aggregate_matrix(
    A: np.ndarray,
    x: np.ndarray,
    va_total: np.ndarray,
    va_comp: np.ndarray,
    sector_map: np.ndarray,
) -> dict:
    """
    Aggregate 402×402 A matrix to 17×17 using output-weighted averaging.

    Each element A_agg[s, t] is the output-weighted average of all
    402-industry A[i, j] pairs where industry i maps to sector s and
    industry j maps to sector t.

    Parameters
    ----------
    A          : (402, 402) domestic technical coefficients
    x          : (402,)     industry gross output ($M)
    va_total   : (402,)     value added ($M)
    va_comp    : (402,)     compensation of employees ($M)
    sector_map : (402,) int sector index for each of 402 industries

    Returns
    -------
    dict with keys:
      A_agg    : np.ndarray (17, 17)  — aggregated technical coefficients
      x_agg    : np.ndarray (17,)     — sector gross output ($M)
      va_share : np.ndarray (17,)     — sector VA / gross output
      li_share : np.ndarray (17,)     — sector compensation / gross output
    """
    A_agg    = np.zeros((N, N), dtype=float)
    x_agg    = np.zeros(N, dtype=float)
    va_agg   = np.zeros(N, dtype=float)
    comp_agg = np.zeros(N, dtype=float)

    for j, s in enumerate(sector_map):
        if s < 0:
            continue
        x_agg[s]    += x[j]
        va_agg[s]   += va_total[j]
        comp_agg[s] += va_comp[j]

    for t in range(N):
        t_cols = np.where(sector_map == t)[0]
        if len(t_cols) == 0:
            continue
        x_t = x[t_cols]
        tot  = x_t.sum()
        if tot <= 0:
            continue
        for s in range(N):
            s_rows = np.where(sector_map == s)[0]
            if len(s_rows) > 0:
                A_agg[s, t] = (
                    A[np.ix_(s_rows, t_cols)].sum(axis=0).dot(x_t) / tot
                )

    with np.errstate(divide="ignore", invalid="ignore"):
        va_share = np.where(x_agg > 0, va_agg   / x_agg, 0.0)
        li_share = np.where(x_agg > 0, comp_agg / x_agg, 0.0)

    return {
        "A_agg":    A_agg,
        "x_agg":    x_agg,
        "va_share": va_share,
        "li_share": li_share,
    }


# ---------------------------------------------------------------------------
# York County income-weighted PCE share vector
# Source: BLS CES 2022 Table 1101 quintile vectors (5 income groups)
#         weighted by York County household income distribution
#         from ACS B19001 2021 5-year estimate (FIPS 51199)
#
# York County quintile weights (vs national 20% each):
#   Q1 Lowest 20%:  8.6%  (York County far fewer low-income households)
#   Q2 Second 20%: 12.2%
#   Q3 Third 20%:  21.7%
#   Q4 Fourth 20%: 28.8%  (York County concentrated in upper-middle income)
#   Q5 Highest 20%:28.8%
#
# Key differences vs national average PCE shares:
#   Education & Health: +4.1pp (higher-income HH spend more on healthcare)
#   Mfg Durable (vehicles): -9.2pp (fewer low-income HH buying used cars)
#   Retail Trade: -3.2pp
# ---------------------------------------------------------------------------

_YORK_PCE_SHARES: np.ndarray = np.array([
    0.000000,  #  0 Agriculture
    0.023772,  #  1 Mining & Utilities
    0.064472,  #  2 Construction
    0.029924,  #  3 Mfg — Durable
    0.257446,  #  4 Mfg — Non-Durable
    0.000000,  #  5 Wholesale Trade
    0.128298,  #  6 Retail Trade
    0.071384,  #  7 Transportation & Whsg.
    0.077021,  #  8 Information
    0.083370,  #  9 Finance & Real Estate
    0.000610,  # 10 Professional Services
    0.000660,  # 11 Admin & Waste Services
    0.041065,  # 12 Education & Health
    0.032101,  # 13 Arts & Accommodation
    0.047225,  # 14 Private Other Services
    0.000000,  # 15 Federal Gov & Defense
    0.142652,  # 16 State & Local Gov
], dtype=float)


def build_pce_shares(
    pce_raw: np.ndarray,
    com_sector: np.ndarray,
    use_york_weights: bool = True,
) -> np.ndarray:
    """
    Return PCE share vector governing household spending in the Type II closure.

    By default returns the York County income-weighted PCE vector derived from:
      - BLS CES 2022 Table 1101 — quintile-specific spending patterns
      - ACS B19001 York County 2021 5-yr — local household income distribution

    When use_york_weights=False, falls back to the national average PCE
    aggregated from the BEA Use Table (previous behaviour).

    Parameters
    ----------
    pce_raw           : (402,) PCE final demand by commodity ($M, Use Table)
    com_sector        : (402,) sector index for each commodity
    use_york_weights  : bool — if True, use York County income-weighted vector

    Returns
    -------
    np.ndarray (17,) — PCE shares summing to 1.0
    """
    if use_york_weights:
        # Verify sum ≈ 1.0 (defensive)
        total = _YORK_PCE_SHARES.sum()
        if abs(total - 1.0) < 0.01:
            return _YORK_PCE_SHARES.copy()
        # If somehow off, renormalise
        return _YORK_PCE_SHARES / total

    # National average fallback — aggregate BEA Use Table PCE by sector
    pce_agg = np.zeros(N, dtype=float)
    for i, s in enumerate(com_sector):
        if 0 <= s < N:
            pce_agg[s] += max(float(pce_raw[i]), 0.0)
    total = pce_agg.sum()
    return pce_agg / total if total > 0 else np.ones(N, dtype=float) / N


# ---------------------------------------------------------------------------
# Regionalization methods
# ---------------------------------------------------------------------------

def compute_sdp_rpc(lq: np.ndarray, alpha: float = 0.20) -> np.ndarray:
    """
    Supply-Demand Pool (SDP) regionalization coefficients.

    Formula: RPC_s = LQ_s / (LQ_s + alpha),  element-wise, capped at 1.0.

    Source
    ------
    Kronenberg (2009) "The Use of Input-Output Analysis in Regional
    Economic Impact Assessment", Review of Regional Research 29(2).
    Flegg & Tohmo (2016) "Estimating Regional Input Coefficients and
    Multipliers", Spatial Economic Analysis 11(4).

    Parameters
    ----------
    lq    : (N,) location quotient array from QCEW
    alpha : float — cross-hauling parameter in (0, 1).
              0.10 = low cross-hauling, dense regional economies
              0.20 = standard for US counties (default)
              0.30 = high cross-hauling, very open small economies

    Returns
    -------
    np.ndarray (N,) — regional purchase coefficients in [0, 1]

    Notes
    -----
    Sectors with LQ=0 receive RPC=0 (no local employment → no local supply).
    RPC is mathematically bounded below 1.0 for any finite LQ and alpha>0;
    the np.minimum cap is retained as a defensive safeguard.
    """
    if not (0 < alpha < 1):
        raise ValueError(f"alpha must be in (0, 1); got {alpha}.")
    if lq.shape != (N,):
        raise ValueError(f"lq must have shape ({N},); got {lq.shape}.")
    with np.errstate(divide="ignore", invalid="ignore"):
        rpc = np.where(lq > 0, lq / (lq + alpha), 0.0)
    return np.minimum(rpc, 1.0)


def compute_flq_rpc(
    lq: np.ndarray,
    total_county_emp: float,
    total_national_emp: float,
    delta: float = 0.25,
) -> tuple[np.ndarray, float]:
    """
    Flegg Location Quotient (FLQ) regionalization — comparison only.

    Not recommended as the primary method for counties with fewer than
    ~100,000 workers; the size penalty collapses to near-zero for small
    counties such as York County VA (~22,000 workers).

    Parameters
    ----------
    lq                 : (N,) location quotient array
    total_county_emp   : float — county total employment (workers)
    total_national_emp : float — national total employment (thousands)
    delta              : float — FLQ sensitivity parameter

    Returns
    -------
    rpc    : np.ndarray (N,) — FLQ regional purchase coefficients
    lambda : float           — FLQ size-adjustment factor
    """
    ratio = total_county_emp / max(total_national_emp, 1.0)
    lam   = float(np.log2(1.0 + ratio) ** delta)
    rpc   = np.minimum(lq * lam, 1.0)
    return rpc, lam


def compute_slq_rpc(lq: np.ndarray) -> np.ndarray:
    """
    Simple Location Quotient (SLQ) — comparison only.

    RPC_s = min(LQ_s, 1.0).

    Parameters
    ----------
    lq : (N,) location quotient array

    Returns
    -------
    np.ndarray (N,) — SLQ regional purchase coefficients
    """
    return np.minimum(lq, 1.0)


def regionalize(
    A_nat17: np.ndarray,
    qcew: dict,
    national_emp_total: float,
    sdp_alpha: float = 0.20,
    flq_delta: float = 0.25,
) -> dict:
    """
    Regionalize the 17×17 national A matrix using SDP as the primary method.

    SLQ and FLQ RPCs are computed for comparison display only and have no
    effect on the model's primary results.

    Parameters
    ----------
    A_nat17           : (17, 17) national domestic technical coefficients
    qcew              : dict from load_qcew containing 'lq' and 'total_county_emp'
    national_emp_total: float — national employment total (thousands)
    sdp_alpha         : float — SDP cross-hauling parameter
    flq_delta         : float — FLQ delta (comparison only)

    Returns
    -------
    dict with keys:
      A_york     : (17, 17) regionalized matrix (SDP primary)
      lq_york    : (17,)    county location quotients
      rpc_york   : (17,)    SDP RPCs (alias for rpc_sdp)
      rpc_sdp    : (17,)    SDP RPCs
      rpc_slq    : (17,)    SLQ RPCs
      rpc_flq    : (17,)    FLQ RPCs
      flq_lambda : float    FLQ size factor
      sdp_alpha  : float
      flq_delta  : float
    """
    lq_york = qcew["lq"].copy()

    rpc_sdp            = compute_sdp_rpc(lq_york, alpha=sdp_alpha)
    rpc_slq            = compute_slq_rpc(lq_york)
    rpc_flq, flq_lam   = compute_flq_rpc(
        lq_york,
        total_county_emp   = qcew["total_county_emp"],
        total_national_emp = national_emp_total * 1000,
        delta              = flq_delta,
    )

    A_york = A_nat17 * rpc_sdp[:, np.newaxis]

    return {
        "A_york":     A_york,
        "lq_york":    lq_york,
        "rpc_york":   rpc_sdp,  # primary; alias retained for backwards compat
        "rpc_sdp":    rpc_sdp,
        "rpc_slq":    rpc_slq,
        "rpc_flq":    rpc_flq,
        "flq_lambda": flq_lam,
        "sdp_alpha":  sdp_alpha,
        "flq_delta":  flq_delta,
    }


# ---------------------------------------------------------------------------
# BLS CES national employment
# ---------------------------------------------------------------------------

# BLS CES series IDs mapped to model sectors.
# Keys are sector indices or sub-sector labels where a sector requires
# multiple series to be summed.
BLS_SERIES: dict[Any, str] = {
    "1a": "CEU1021000001",   # Mining
    "1b": "CEU0622000001",   # Utilities (combined with mining → sector 1)
    2:    "CEU2000000001",   # Construction
    3:    "CEU3100000001",   # Durable manufacturing
    4:    "CEU3200000001",   # Non-durable manufacturing
    5:    "CEU4142000001",   # Wholesale trade
    6:    "CEU4200000001",   # Retail trade
    7:    "CEU4300000001",   # Transportation & warehousing
    8:    "CEU5000000001",   # Information
    "9a": "CEU5552000001",   # Finance
    "9b": "CEU5553000001",   # Real estate
    "10a":"CEU6054000001",   # Professional & technical services
    "10b":"CEU6055000001",   # Management of companies
    11:   "CEU6056000001",   # Admin & waste services
    "12a":"CEU6561000001",   # Education services
    "12b":"CEU6562000001",   # Health care & social assistance
    "13a":"CEU7071000001",   # Arts & entertainment
    "13b":"CEU7072000001",   # Accommodation & food services
    14:   "CEU8000000001",   # Private other services (NAICS 81)
    15:   "CEU9091000001",   # Federal government
    "16a":"CEU9092000001",   # State government
    "16b":"CEU9093000001",   # Local government
}
ALL_SERIES_IDS: list[str] = list(BLS_SERIES.values())

# Agriculture national employment (thousands of workers, BLS/USDA 2022 annual avg).
# BLS CES does not cover farm workers; this constant is standard practice.
# Source: USDA/BLS published figure consistent with I-O literature.
AGR_NATIONAL_EMP_FALLBACK: float = 2109.9

# BLS CES 2022 annual average employment (thousands of workers).
# Used when the live API call fails.
# Source: BLS CES published 2022 annual averages.
BLS_FALLBACK_2022: dict[str, float] = {
    "CEU1021000001":  657.5,   # Mining
    "CEU0622000001":  574.0,   # Utilities
    "CEU2000000001": 7697.2,   # Construction
    "CEU3100000001": 9122.5,   # Durable mfg
    "CEU3200000001": 4841.6,   # Non-durable mfg
    "CEU4142000001": 6153.4,   # Wholesale
    "CEU4200000001":15592.8,   # Retail
    "CEU4300000001": 6622.7,   # Transportation
    "CEU5000000001": 3038.9,   # Information
    "CEU5552000001": 6617.6,   # Finance
    "CEU5553000001": 2394.6,   # Real estate
    "CEU6054000001":10199.4,   # Professional
    "CEU6055000001": 2906.5,   # Management
    "CEU6056000001": 9628.6,   # Admin
    "CEU6561000001": 3787.0,   # Education
    "CEU6562000001":20411.9,   # Health
    "CEU7071000001": 2615.7,   # Arts
    "CEU7072000001":12346.2,   # Accommodation
    "CEU8000000001": 5697.4,   # Private other services
    "CEU9091000001": 2997.0,   # Federal government
    "CEU9092000001": 5220.0,   # State government
    "CEU9093000001":14307.2,   # Local government
}


def fetch_bls(api_key: str, year: str = "2022") -> dict[str, float] | None:
    """
    Fetch annual average employment from BLS CES for all model sectors.

    Parameters
    ----------
    api_key : str — BLS API registration key
    year    : str — data year (annual average)

    Returns
    -------
    dict mapping series ID → employment (thousands of workers), or None
    on failure.
    """
    payload = {
        "seriesid":       ALL_SERIES_IDS,
        "startyear":      year,
        "endyear":        year,
        "annualaverage":  "true",
        "registrationkey": api_key,
    }
    try:
        r = requests.post(BLS_BASE, json=payload, timeout=45)
        r.raise_for_status()
        resp = r.json()
        if resp.get("status") != "REQUEST_SUCCEEDED":
            return None
        raw: dict[str, float] = {}
        for series in resp["Results"]["series"]:
            sid = series["seriesID"]
            # Prefer the annual average (period M13); fall back to mean of monthlies
            val = next(
                (float(d["value"]) for d in series["data"] if d["period"] == "M13"),
                None,
            )
            if val is None:
                monthly = [
                    float(d["value"])
                    for d in series["data"]
                    if d["period"].startswith("M") and d["period"] != "M13"
                ]
                val = float(np.mean(monthly)) if monthly else 0.0
            raw[sid] = val
        return raw
    except Exception:
        return None


def series_to_emp(raw: dict[str, float]) -> np.ndarray:
    """
    Convert a BLS CES series dict to a (17,) employment array.

    Agriculture is set from AGR_NATIONAL_EMP_FALLBACK (BLS CES does not
    cover farm workers). Sectors requiring multiple series are summed.

    Parameters
    ----------
    raw : dict mapping BLS series ID → employment (thousands of workers)

    Returns
    -------
    np.ndarray (17,) — annual average employment by sector (thousands of workers)
    """
    emp = np.zeros(N, dtype=float)
    emp[0] = AGR_NATIONAL_EMP_FALLBACK  # Agriculture: constant (see note above)

    # Single-series sectors
    for idx in (2, 3, 4, 5, 6, 7, 8, 11, 14):
        emp[idx] = raw.get(BLS_SERIES[idx], 0.0)

    # Multi-series sectors (sum sub-components)
    emp[1]  = raw.get(BLS_SERIES["1a"],  0.0) + raw.get(BLS_SERIES["1b"],  0.0)
    emp[9]  = raw.get(BLS_SERIES["9a"],  0.0) + raw.get(BLS_SERIES["9b"],  0.0)
    emp[10] = raw.get(BLS_SERIES["10a"], 0.0) + raw.get(BLS_SERIES["10b"], 0.0)
    emp[12] = raw.get(BLS_SERIES["12a"], 0.0) + raw.get(BLS_SERIES["12b"], 0.0)
    emp[13] = raw.get(BLS_SERIES["13a"], 0.0) + raw.get(BLS_SERIES["13b"], 0.0)
    emp[15] = raw.get(BLS_SERIES[15],    0.0)
    emp[16] = raw.get(BLS_SERIES["16a"], 0.0) + raw.get(BLS_SERIES["16b"], 0.0)

    return emp


# ---------------------------------------------------------------------------
# BEA GDP-by-Industry API
# ---------------------------------------------------------------------------

def _bea_extract_rows(data: dict) -> list[dict]:
    """
    Extract the 'Data' list from a BEA API JSON response.

    Handles both dict and list formats of the 'Results' field, which
    BEA has changed across API versions.
    """
    results = data.get("BEAAPI", {}).get("Results", {})
    if isinstance(results, list):
        for item in results:
            if isinstance(item, dict) and "Data" in item:
                return item["Data"]
        return []
    if isinstance(results, dict):
        return results.get("Data", [])
    return []


def fetch_bea_table(
    api_key: str,
    table_id: str,
    year: str = "2022",
) -> tuple[np.ndarray | None, str]:
    """
    Fetch a BEA GDPbyIndustry table and aggregate to 17 model sectors.

    BEA table IDs (verified 2024-2025):
      "VA" → TableID 1  — Value Added by Industry   (units: $B, converted to $M)
      "GO" → TableID 15 — Gross Output by Industry  (units: $B, converted to $M)

    BEA changed units from $M to $B sometime in 2023-2024. All DataValues
    are multiplied by 1000 to restore $M for consistency with the fallback
    arrays and downstream calculations.

    Parameters
    ----------
    api_key  : str — BEA API key
    table_id : str — logical table identifier: "VA" or "GO"
    year     : str — data year

    Returns
    -------
    (result, diagnostic_string)
      result is a (17,) array ($M) on success, or None on failure.
      diagnostic_string describes rows matched, total, or error.
    """
    TABLE_MAP: dict[str, str] = {"VA": "1", "GO": "15"}
    bea_tid = TABLE_MAP.get(table_id, table_id)

    params = {
        "UserID":      api_key,
        "method":      "GetData",
        "datasetname": "GDPbyIndustry",
        "TableID":     bea_tid,
        "Frequency":   "A",
        "Year":        year,
        "Industry":    "ALL",
        "ResultFormat":"JSON",
    }
    try:
        r = requests.get(BEA_BASE, params=params, timeout=30)
        r.raise_for_status()
        rows = _bea_extract_rows(r.json())
        if not rows:
            return None, f"TableID={bea_tid}: API returned 0 data rows."
        result = np.zeros(N, dtype=float)
        matched = 0
        for row in rows:
            ind = str(row.get("Industry", "")).strip()
            if ind in BEA_SUMMARY_TO_SECTOR:
                try:
                    val_b = float(str(row.get("DataValue", "")).replace(",", ""))
                    result[BEA_SUMMARY_TO_SECTOR[ind]] += val_b * 1000.0  # $B → $M
                    matched += 1
                except ValueError:
                    pass
        total_t = result.sum() / 1e6  # $M → $T for display
        diag = (
            f"TableID={bea_tid} rows={len(rows)} "
            f"matched={matched} sum=${total_t:.1f}T"
        )
        return result, diag
    except Exception as exc:
        return None, f"TableID={bea_tid} EXCEPTION: {exc}"


def fetch_bea_nipa_mpc(api_key: str, year: str = "2022") -> float | None:
    """
    Fetch the marginal propensity to consume (MPC) from BEA NIPA Table T20600.

    MPC = Personal Outlays (A068RC) / Disposable Personal Income (A065RC).

    Parameters
    ----------
    api_key : str — BEA API key
    year    : str — data year

    Returns
    -------
    float — MPC estimate, or None if the API call fails or series are absent.
    """
    params = {
        "UserID":      api_key,
        "method":      "GetData",
        "datasetname": "NIPA",
        "TableName":   "T20600",
        "Frequency":   "A",
        "Year":        year,
        "ResultFormat":"JSON",
    }
    try:
        r = requests.get(BEA_BASE, params=params, timeout=30)
        r.raise_for_status()
        vals: dict[str, float] = {}
        for row in _bea_extract_rows(r.json()):
            sc = row.get("SeriesCode", "")
            try:
                vals[sc] = float(str(row.get("DataValue", "")).replace(",", ""))
            except ValueError:
                pass
        dpi     = vals.get("A065RC")  # Disposable Personal Income ($B)
        outlays = vals.get("A068RC")  # Personal Outlays ($B)
        if dpi and outlays and dpi > 0:
            return outlays / dpi
        return None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Data source fallback logic
# ---------------------------------------------------------------------------

def _load_csv_cache(filename: str) -> np.ndarray | None:
    """
    Load a sector-level BEA data array from the local CSV cache.

    CSV must contain columns 'sector' and 'value_millions'. All 17
    SECTOR_LABELS must be present for the result to be used.

    Parameters
    ----------
    filename : str — filename within DATA_DIR (e.g. 'bea_value_added.csv')

    Returns
    -------
    np.ndarray (17,) in $M, or None if the file is missing, malformed,
    or contains zero total.
    """
    path = DATA_DIR / filename
    if not path.exists():
        return None
    try:
        df = pd.read_csv(path)
        if "sector" not in df.columns or "value_millions" not in df.columns:
            warnings.warn(
                f"CSV cache '{filename}' missing 'sector' or "
                "'value_millions' column. File will be ignored.",
                UserWarning,
                stacklevel=2,
            )
            return None
        result = np.zeros(N, dtype=float)
        for i, label in enumerate(SECTOR_LABELS):
            match = df[df["sector"] == label]["value_millions"]
            if len(match) > 0:
                result[i] = float(match.values[0])
        return result if result.sum() > 0 else None
    except Exception:
        return None


def _use_or_fallback(
    result_and_diag: tuple[np.ndarray | None, str] | np.ndarray | None,
    api_fallback: np.ndarray,
    csv_filename: str,
    min_plausible_sum: float | None = None,
) -> tuple[np.ndarray, str, str]:
    """
    Three-tier data source selection with full diagnostic logging.

    Priority:
      1. CSV cache (primary) — versioned, stable, user-controlled.
      2. Live API result     — used only if CSV is absent and API succeeded.
      3. Hardcoded constant  — last resort.

    Parameters
    ----------
    result_and_diag   : (array, diag_string) tuple from fetch_bea_table,
                        or a bare array, or None.
    api_fallback      : (17,) hardcoded fallback array ($M)
    csv_filename      : filename within DATA_DIR
    min_plausible_sum : if set, the API result must sum to at least this
                        value ($M) to be considered plausible.

    Returns
    -------
    (array, source_label, diagnostic_string)
      source_label : one of 'csv_cache', 'live', 'hardcoded_fallback'
    """
    if isinstance(result_and_diag, tuple):
        api_result, api_diag = result_and_diag
    else:
        api_result, api_diag = result_and_diag, ""

    csv_result = _load_csv_cache(csv_filename)

    # Tier 1: CSV cache
    if csv_result is not None:
        api_ok = (
            api_result is not None
            and api_result.sum() > 0
            and (min_plausible_sum is None or api_result.sum() >= min_plausible_sum)
        )
        note = (
            f"API also available ({api_diag}) — CSV used as primary."
            if api_ok
            else f"API unavailable or implausible ({api_diag}) — CSV used."
        )
        return csv_result, "csv_cache", note

    # Tier 2: live API
    if api_result is not None and api_result.sum() > 0:
        if min_plausible_sum is None or api_result.sum() >= min_plausible_sum:
            return api_result, "live", f"CSV not found; using live API. {api_diag}"

    # Tier 3: hardcoded constant
    return (
        api_fallback,
        "hardcoded_fallback",
        f"CSV not found; API failed. Using 2022 constant. {api_diag}",
    )


# ---------------------------------------------------------------------------
# Employment coefficients
# ---------------------------------------------------------------------------

def build_employment_coefficients(
    national_emp: np.ndarray,
    bea_output: np.ndarray,
    bea_comp: np.ndarray,
    bea_va: np.ndarray,
    qcew: dict,
    va_share_agg: np.ndarray,
    mpc: float = 0.90,
    regional_retention: float = 0.50,
) -> dict:
    """
    Compute sector-level employment and income coefficients.

    Employment intensity
    -------------------
    jobs_per_va[s] = national_emp[s] (000s of workers) × 1000
                     / bea_va[s] ($M value added)
    Units: workers per $1M value added.

    Wage calibration
    ----------------
    Uses York County average annual wages from QCEW where the county
    has positive employment; falls back to the national average otherwise.

    Labor income share
    ------------------
    york_li_share[s] = (york_wage[s] × national_emp[s] × 1000)
                       / (bea_output[s] × 1e6)
    Clipped to [0, 0.70] to exclude implausible values.

    Household spending share
    ------------------------
    hh_share = mpc × regional_retention
    Share of labor income that re-enters the county economy as household
    demand.

    Parameters
    ----------
    national_emp       : (17,) BLS CES employment (thousands of workers)
    bea_output         : (17,) BEA gross output ($M)
    bea_comp           : (17,) BEA compensation of employees ($M)
    bea_va             : (17,) BEA value added ($M)
    qcew               : dict from load_qcew
    va_share_agg       : (17,) VA/output shares from aggregate_matrix
    mpc                : float — marginal propensity to consume (BEA NIPA)
    regional_retention : float — fraction of household spending staying local

    Returns
    -------
    dict with keys:
      jobs_per_va        : (17,) workers per $1M value added
      jobs_per_million   : alias for jobs_per_va (backwards compat)
      avg_wage           : (17,) average annual pay ($ per worker)
      li_share           : (17,) labour income / gross output
      va_share           : (17,) value added / gross output
      hh_share           : float — mpc × regional_retention
      nat_avg_wage       : (17,) national average annual pay ($)
      mpc                : float
      regional_retention : float
    """
    with np.errstate(divide="ignore", invalid="ignore"):
        jobs_per_va  = np.where(bea_va > 0, (national_emp * 1000) / bea_va, 0.0)
        nat_avg_wage = np.where(
            national_emp > 0,
            (bea_comp * 1e6) / (national_emp * 1000),
            0.0,
        )
        li_share_nat = np.where(bea_output > 0, bea_comp / bea_output, 0.0)

    # Use county wage where QCEW reports local employment; else national average
    york_wage = np.where(qcew["emp"] > 0, qcew["wage"], nat_avg_wage)

    york_li_share = np.where(
        bea_output > 0,
        (york_wage * national_emp * 1000) / (bea_output * 1e6),
        li_share_nat,
    )
    york_li_share = np.clip(york_li_share, 0.0, 0.70)

    hh_share = mpc * regional_retention

    return {
        "jobs_per_va":        jobs_per_va,
        "jobs_per_million":   jobs_per_va,   # backwards-compat alias
        "avg_wage":           york_wage,
        "li_share":           york_li_share,
        "va_share":           va_share_agg,
        "hh_share":           hh_share,
        "nat_avg_wage":       nat_avg_wage,
        "mpc":                mpc,
        "regional_retention": regional_retention,
    }


# ---------------------------------------------------------------------------
# Leontief engine
# ---------------------------------------------------------------------------

def check_leontief_stability(A: np.ndarray) -> dict:
    """
    Compute stability diagnostics for the regionalized A matrix.

    A Leontief system is stable (inverse converges) if and only if the
    spectral radius ρ(A) < 1.0.

    Parameters
    ----------
    A : (N, N) technical coefficients matrix

    Returns
    -------
    dict with keys:
      spectral_radius : float  — max absolute eigenvalue
      stable          : bool   — True if ρ < 1.0
      max_col_sum     : float  — max column sum of A
      col_sums        : (N,)   — column sums
      warnings        : list[str] — human-readable warnings if ρ ≥ 0.95
    """
    eigenvalues     = np.linalg.eigvals(A)
    spectral_radius = float(np.max(np.abs(eigenvalues)))
    col_sums        = A.sum(axis=0)
    warn_list: list[str] = []

    if spectral_radius >= 1.0:
        warn_list.append(
            f"CRITICAL: spectral radius ρ={spectral_radius:.4f} ≥ 1.0. "
            "Leontief inverse will not converge."
        )
    elif spectral_radius > 0.95:
        warn_list.append(
            f"WARNING: spectral radius ρ={spectral_radius:.4f} is close to 1.0. "
            "Results may be numerically sensitive."
        )
    if col_sums.max() >= 1.0:
        n_bad = int((col_sums >= 1.0).sum())
        warn_list.append(
            f"WARNING: {n_bad} column(s) of A sum to ≥ 1.0. "
            "These were rescaled during build_A_domestic."
        )

    return {
        "spectral_radius": spectral_radius,
        "stable":          spectral_radius < 1.0,
        "max_col_sum":     float(col_sums.max()),
        "col_sums":        col_sums,
        "warnings":        warn_list,
    }


def check_multipliers(mults: dict) -> list[dict]:
    """
    Validate output multipliers against published plausibility ranges.

    Ranges sourced from BEA RIMS II User Handbook (2022) and IMPLAN
    validation studies for US counties.

    Parameters
    ----------
    mults : dict from compute_impacts containing 'type1', 'type2', 'emp'

    Returns
    -------
    list of dicts with keys: metric, value, low, high, ok, note
    """
    return [
        {
            "metric": "Type I Output Multiplier",
            "value":  mults["type1"],
            "low":    MULT_TYPE1_MIN,
            "high":   MULT_TYPE1_MAX,
            "ok":     MULT_TYPE1_MIN <= mults["type1"] <= MULT_TYPE1_MAX,
            "note":   "Indirect supply-chain effects only",
        },
        {
            "metric": "Type II Output Multiplier",
            "value":  mults["type2"],
            "low":    MULT_TYPE2_MIN,
            "high":   MULT_TYPE2_MAX,
            "ok":     MULT_TYPE2_MIN <= mults["type2"] <= MULT_TYPE2_MAX,
            "note":   "Includes induced household spending",
        },
        {
            "metric": "Employment Multiplier",
            "value":  mults["emp"],
            "low":    1.1,
            "high":   4.0,
            "ok":     1.1 <= mults["emp"] <= 4.0,
            "note":   "Total jobs / direct jobs",
        },
    ]


def build_leontief_inverses(
    A_reg: np.ndarray,
    li_share: np.ndarray,
    pce_shares: np.ndarray,
    hh_share: float,
) -> dict:
    """
    Pre-compute Type I (L1) and Type II (L2_sub) Leontief inverses.

    L1     = (I − A_reg)^{-1}
    L2_sub = upper-left N×N block of (I − A2)^{-1}
    where A2 is the (N+1)×(N+1) household-closed matrix appending a
    household row and column.

    A near-singular check is performed before inversion. If the spectral
    radius is ≥ 1.0 a ValueError is raised rather than returning a
    numerically meaningless inverse.

    Parameters
    ----------
    A_reg      : (N, N) regionalized technical coefficients (SDP-adjusted)
    li_share   : (N,) labour income share (comp / gross output)
    pce_shares : (N,) household spending shares by sector
    hh_share   : float — mpc × regional_retention

    Returns
    -------
    dict with keys:
      L1        : (N, N) Type I Leontief inverse
      L2_sub    : (N, N) Type II induced-effects sub-matrix
      stability : dict from check_leontief_stability
    """
    stability = check_leontief_stability(A_reg)
    if not stability["stable"]:
        raise ValueError(
            f"Cannot invert: spectral radius ρ={stability['spectral_radius']:.4f} ≥ 1.0. "
            "Check that input files are the correct BEA benchmark tables."
        )

    n  = len(A_reg)
    I  = np.eye(n)
    L1 = np.linalg.inv(I - A_reg)

    # Household-closed augmented matrix for Type II
    A2 = np.zeros((n + 1, n + 1), dtype=float)
    A2[:n, :n] = A_reg
    A2[:n,  n] = pce_shares * hh_share  # household demand column
    A2[ n, :n] = li_share               # household income row
    L2_sub = np.linalg.inv(np.eye(n + 1) - A2)[:n, :n]

    return {"L1": L1, "L2_sub": L2_sub, "stability": stability}


def compute_impacts(
    Y: np.ndarray,
    inverses: dict,
    emp_coeffs: dict,
) -> dict:
    """
    Compute direct, indirect, induced, and total economic impacts.

    Direct effects:   the initial demand vector Y ($M by sector)
    Indirect effects: (L1 − I) @ Y  — supply-chain response ($M)
    Induced effects:  (L2_sub − L1) @ Y  — household spending loop ($M)
    Total:            direct + indirect + induced

    Negative induced values (numerical artefacts) are floored to zero.

    Parameters
    ----------
    Y         : (N,) final demand vector ($M)
    inverses  : dict from build_leontief_inverses
    emp_coeffs: dict from build_employment_coefficients

    Returns
    -------
    dict with keys:
      direct, indirect, induced, total : impact dicts (see _impact_dict)
      multipliers   : dict — type1, type2, emp, li, va multipliers
      mult_checks   : list from check_multipliers
      vecs          : dict — raw sector vectors for direct/indirect/induced/total
      diagnostics   : dict — pct share of each effect type in total output
    """
    L1     = inverses["L1"]
    L2_sub = inverses["L2_sub"]
    I_n    = np.eye(len(Y))

    direct   = Y.copy()
    indirect = np.maximum((L1 - I_n) @ Y, 0.0)
    induced  = np.maximum((L2_sub - L1) @ Y, 0.0)
    total    = direct + indirect + induced

    def _impact_dict(vec: np.ndarray) -> dict:
        """Scalar aggregates for a single effect-type output vector."""
        vm = vec / 1e6  # $M → $M/million for jobs-per-$M arithmetic
        return {
            "output":       float(vec.sum()),          # $
            "jobs":         float((vm * emp_coeffs["jobs_per_va"]).sum()),
            "labor_income": float((vec * emp_coeffs["li_share"]).sum()),  # $
            "value_added":  float((vec * emp_coeffs["va_share"]).sum()),  # $
            "output_vec":   vec.copy(),
            "jobs_vec":     (vm * emp_coeffs["jobs_per_va"]).copy(),
            "li_vec":       (vec * emp_coeffs["li_share"]).copy(),
            "va_vec":       (vec * emp_coeffs["va_share"]).copy(),
        }

    d, ii, ind, t = (
        _impact_dict(direct),
        _impact_dict(indirect),
        _impact_dict(induced),
        _impact_dict(total),
    )

    def _mult(numerator: float, denominator: float) -> float:
        return numerator / denominator if denominator > 0 else 0.0

    mults = {
        "type1": _mult(d["output"] + ii["output"],  d["output"]),
        "type2": _mult(t["output"],                 d["output"]),
        "emp":   _mult(t["jobs"],   max(d["jobs"],  1e-9)),
        "li":    _mult(t["labor_income"], max(d["labor_income"], 1e-9)),
        "va":    _mult(t["value_added"],  max(d["value_added"],  1e-9)),
    }

    total_out = max(t["output"], 1.0)  # avoid div-by-zero in diagnostics
    return {
        "direct":      d,
        "indirect":    ii,
        "induced":     ind,
        "total":       t,
        "multipliers": mults,
        "mult_checks": check_multipliers(mults),
        "vecs": {
            "direct":   direct,
            "indirect": indirect,
            "induced":  induced,
            "total":    total,
        },
        "diagnostics": {
            "direct_pct":   d["output"]  / total_out * 100,
            "indirect_pct": ii["output"] / total_out * 100,
            "induced_pct":  ind["output"]/ total_out * 100,
        },
    }


# ---------------------------------------------------------------------------
# Uncertainty bands
# ---------------------------------------------------------------------------

def add_uncertainty(results: dict, pct: float = UNCERTAINTY_PCT) -> dict:
    """
    Append symmetric ±pct uncertainty bands to all scalar impact outputs.

    Basis: Watson et al. (2015) Journal of Regional Science 55(1) documented
    ±20–30% gaps between Leontief I-O model predictions and observed economic
    outcomes. An additional ±10–15% is included for applying 2017 production
    technology to the current-year economy.

    Parameters
    ----------
    results : dict from compute_impacts
    pct     : float — fractional uncertainty (default 0.30 = ±30%)

    Returns
    -------
    Updated results dict with a 'bands' sub-dict added to each impact type.
    The 'uncertainty_pct' key records the applied percentage.
    """
    if not (0 < pct < 1):
        raise ValueError(f"pct must be in (0, 1); got {pct}.")

    out: dict = {}
    for key in ("direct", "indirect", "induced", "total"):
        imp = results[key]
        out[key] = {
            **imp,
            "bands": {
                "output_low":        imp["output"]       * (1 - pct),
                "output_high":       imp["output"]       * (1 + pct),
                "jobs_low":          imp["jobs"]         * (1 - pct),
                "jobs_high":         imp["jobs"]         * (1 + pct),
                "labor_income_low":  imp["labor_income"] * (1 - pct),
                "labor_income_high": imp["labor_income"] * (1 + pct),
                "value_added_low":   imp["value_added"]  * (1 - pct),
                "value_added_high":  imp["value_added"]  * (1 + pct),
            },
        }
    out["multipliers"]     = results["multipliers"]
    out["mult_checks"]     = results["mult_checks"]
    out["vecs"]            = results["vecs"]
    out["diagnostics"]     = results["diagnostics"]
    out["uncertainty_pct"] = pct
    return out


# ---------------------------------------------------------------------------
# Sensitivity analysis
# ---------------------------------------------------------------------------

def sensitivity_analysis(
    Y: np.ndarray,
    A_nat17: np.ndarray,
    qcew: dict,
    national_emp_total: float,
    emp_coeffs: dict,
    pce_shares: np.ndarray,
    alpha_range: tuple[float, ...] = (0.10, 0.15, 0.20, 0.25, 0.30),
) -> list[dict]:
    """
    Run impact model across a range of SDP alpha values and comparison methods.

    Returns one row per alpha value plus rows for SLQ and FLQ (delta=0.25)
    for reference. All rows use the same Y, A_nat17, and emp_coeffs.

    Parameters
    ----------
    Y                   : (N,) demand vector ($M)
    A_nat17             : (N, N) national 17-sector A matrix
    qcew                : dict from load_qcew
    national_emp_total  : float — national employment total (thousands)
    emp_coeffs          : dict from build_employment_coefficients
    pce_shares          : (N,) PCE shares
    alpha_range         : SDP alpha values to sweep

    Returns
    -------
    list of dicts, each with keys:
      method, alpha, avg_rpc, total_jobs, total_output,
      type1, type2, emp_mult
    """
    rows: list[dict] = []

    for alpha in alpha_range:
        rpc_sdp = compute_sdp_rpc(qcew["lq"], alpha=alpha)
        A_york  = A_nat17 * rpc_sdp[:, np.newaxis]
        inv     = build_leontief_inverses(
            A_york, emp_coeffs["li_share"], pce_shares, emp_coeffs["hh_share"]
        )
        res = compute_impacts(Y, inv, emp_coeffs)
        rows.append({
            "method":       f"SDP α={alpha:.2f}",
            "alpha":        alpha,
            "avg_rpc":      round(float(rpc_sdp.mean()), 4),
            "total_jobs":   round(res["total"]["jobs"], 1),
            "total_output": round(res["total"]["output"], 0),
            "type1":        round(res["multipliers"]["type1"], 4),
            "type2":        round(res["multipliers"]["type2"], 4),
            "emp_mult":     round(res["multipliers"]["emp"],   4),
        })

    # SLQ reference row
    rpc_slq = compute_slq_rpc(qcew["lq"])
    inv_slq = build_leontief_inverses(
        A_nat17 * rpc_slq[:, np.newaxis],
        emp_coeffs["li_share"], pce_shares, emp_coeffs["hh_share"],
    )
    res_slq = compute_impacts(Y, inv_slq, emp_coeffs)
    rows.append({
        "method":       "SLQ (simple LQ)",
        "alpha":        None,
        "avg_rpc":      round(float(rpc_slq.mean()), 4),
        "total_jobs":   round(res_slq["total"]["jobs"], 1),
        "total_output": round(res_slq["total"]["output"], 0),
        "type1":        round(res_slq["multipliers"]["type1"], 4),
        "type2":        round(res_slq["multipliers"]["type2"], 4),
        "emp_mult":     round(res_slq["multipliers"]["emp"],   4),
    })

    # FLQ reference row (delta=0.25)
    rpc_flq, _ = compute_flq_rpc(
        qcew["lq"],
        qcew["total_county_emp"],
        national_emp_total * 1000,
        delta=0.25,
    )
    inv_flq = build_leontief_inverses(
        A_nat17 * rpc_flq[:, np.newaxis],
        emp_coeffs["li_share"], pce_shares, emp_coeffs["hh_share"],
    )
    res_flq = compute_impacts(Y, inv_flq, emp_coeffs)
    rows.append({
        "method":       "FLQ δ=0.25 (reference)",
        "alpha":        None,
        "avg_rpc":      round(float(rpc_flq.mean()), 4),
        "total_jobs":   round(res_flq["total"]["jobs"], 1),
        "total_output": round(res_flq["total"]["output"], 0),
        "type1":        round(res_flq["multipliers"]["type1"], 4),
        "type2":        round(res_flq["multipliers"]["type2"], 4),
        "emp_mult":     round(res_flq["multipliers"]["emp"],   4),
    })

    return rows


# ---------------------------------------------------------------------------
# Spending profiles
# ---------------------------------------------------------------------------

def get_spending_profile(A_nat17: np.ndarray, sector_idx: int) -> dict[int, float]:
    """
    Return the normalised supply-chain spending profile for one sector.

    Parameters
    ----------
    A_nat17    : (N, N) national technical coefficients
    sector_idx : int — column index (purchasing sector)

    Returns
    -------
    dict mapping sector_index → share of total column sum.
    Only shares above 0.001 are returned. If the column is all zeros,
    returns {sector_idx: 1.0} (self-supply).
    """
    if not (0 <= sector_idx < N):
        raise ValueError(f"sector_idx must be in [0, {N-1}]; got {sector_idx}.")
    col   = A_nat17[:, sector_idx].copy()
    total = col.sum()
    if total <= 0:
        return {sector_idx: 1.0}
    profile = col / total
    return {s: float(v) for s, v in enumerate(profile) if v > 0.001}


def build_all_profiles(A_nat17: np.ndarray) -> dict[int, dict[int, float]]:
    """Build spending profiles for all 17 sectors. See get_spending_profile."""
    return {s: get_spending_profile(A_nat17, s) for s in range(N)}


# ---------------------------------------------------------------------------
# Coefficient validation
# ---------------------------------------------------------------------------

def validate_coefficients(emp_coeffs: dict) -> list[dict]:
    """
    Check sector employment coefficients against published plausibility ranges.

    Ranges:
      jobs/$1M VA : [0.5, 40.0]  — consistent with BLS/BEA cross-sector data
      avg_wage    : [$15,000, $200,000]
      li_share    : [5%, 70%]

    Parameters
    ----------
    emp_coeffs : dict from build_employment_coefficients

    Returns
    -------
    list of dicts with keys: sector, jpm, wage, li, ok, flags
    """
    results: list[dict] = []
    for s in range(N):
        jpm   = float(emp_coeffs["jobs_per_va"][s])
        wage  = float(emp_coeffs["avg_wage"][s])
        li    = float(emp_coeffs["li_share"][s])
        flags: list[str] = []
        if not (0.5 <= jpm <= 40.0):
            flags.append(f"jobs/$1M VA={jpm:.2f} outside [0.5, 40.0]")
        if not (15_000 <= wage <= 200_000):
            flags.append(f"avg_wage=${wage:,.0f} outside [$15k, $200k]")
        if not (0.05 <= li <= 0.70):
            flags.append(f"li_share={li:.1%} outside [5%, 70%]")
        results.append({
            "sector": SECTOR_LABELS[s],
            "jpm":    jpm,
            "wage":   wage,
            "li":     li,
            "ok":     len(flags) == 0,
            "flags":  flags,
        })
    return results


# ---------------------------------------------------------------------------
# Scenario store
# ---------------------------------------------------------------------------

def make_scenario(
    name: str,
    naics: str | int,
    investment: float,
    results_with_bands: dict,
    run_params: dict,
) -> dict:
    """
    Package a completed run into a storable scenario record.

    Parameters
    ----------
    name               : human-readable project name
    naics              : NAICS code used for the run
    investment         : total investment amount ($)
    results_with_bands : dict from add_uncertainty
    run_params         : arbitrary dict of run parameters (stored verbatim)

    Returns
    -------
    dict suitable for appending to a scenario comparison list
    """
    t = results_with_bands["total"]
    m = results_with_bands["multipliers"]
    return {
        "name":         name,
        "naics":        str(naics),
        "investment":   float(investment),
        "sector":       SECTOR_LABELS[naics_to_sector(str(naics))],
        "output":       t["output"],
        "jobs":         t["jobs"],
        "jobs_low":     t["bands"]["jobs_low"],
        "jobs_high":    t["bands"]["jobs_high"],
        "labor_income": t["labor_income"],
        "value_added":  t["value_added"],
        "type1":        m["type1"],
        "type2":        m["type2"],
        "emp_mult":     m["emp"],
        "params":       run_params,
        "full_results": results_with_bands,
    }


# ---------------------------------------------------------------------------
# Full build pipeline
# ---------------------------------------------------------------------------

def build_model(
    use_file: Any,
    ms_file: Any,
    dom_file: Any,
    qcew_file: Any,
    bea_api_key: str,
    bls_api_key: str,
    sdp_alpha: float = 0.20,
    flq_delta: float = 0.25,
    mpc: float = 0.90,
    regional_retention: float = 0.50,
    progress_callback: Any = None,
) -> dict:
    """
    End-to-end model build pipeline.

    Parses all input files, fetches live API data, constructs the
    regionalized Leontief model, and pre-computes inverses and spending
    profiles for repeated use across scenarios.

    Parameters
    ----------
    use_file           : BEA Use Table workbook (file-like or path)
    ms_file            : BEA Market Share Matrix D workbook
    dom_file           : BEA Domestic Direct Requirements B workbook
    qcew_file          : BLS QCEW county CSV
    bea_api_key        : str — BEA API key
    bls_api_key        : str — BLS API key
    sdp_alpha          : float — SDP cross-hauling parameter (default 0.20)
    flq_delta          : float — FLQ delta for comparison display (default 0.25)
    mpc                : float — fallback MPC if BEA NIPA API fails
    regional_retention : float — fraction of household spending remaining local
    progress_callback  : callable(msg: str, pct: int) → None, or None

    Returns
    -------
    dict containing all model state needed for scenario runs:
      A_nat17, A_york, lq_york, rpc_york/rpc_sdp/rpc_slq/rpc_flq,
      flq_lambda, sdp_alpha, flq_delta, pce_shares, inverses, emp_coeffs,
      national_emp, bea_output, bea_comp, bea_va, qcew, all_profiles,
      validation, nat_mpc, regional_retention,
      bea_output_src/diag, bea_comp_src, bea_va_src/diag,
      bls_src (live/fallback per series),
      import_removed_mean, import_leakage_rate (SDP/SLQ/FLQ),
      x_sum, ind_sector, com_sector, stability,
      A_nat15 (backwards-compat alias for A_nat17), scenarios (empty list)
    """
    def _prog(msg: str, pct: int) -> None:
        if progress_callback is not None:
            progress_callback(msg, pct)

    # ── Parse BEA structural tables ──────────────────────────────────────────
    _prog("Parsing BEA Use Table…", 5)
    use_data = parse_use_table(use_file)

    _prog("Parsing Market Share Matrix D…", 13)
    D_data = parse_D_matrix(ms_file)

    _prog("Parsing Domestic Direct Requirements B…", 22)
    B_dom   = parse_B_domestic(dom_file)
    B_total = build_B_total(use_data)

    _prog("Building A_domestic = D @ B_domestic…", 32)
    A_data = build_A_domestic(D_data["D"], B_dom, use_data)

    # ── Sector crosswalk: 402 → 17 ───────────────────────────────────────────
    _prog("Computing sector crosswalk (402 → 17)…", 40)
    ind_sector = np.array(
        [_code_to_sector(c) for c in D_data["ind_codes"]], dtype=int
    )
    com_sector = np.array(
        [_code_to_sector(c) for c in use_data["com_codes"]], dtype=int
    )

    # ── Aggregate to 17-sector matrices ──────────────────────────────────────
    _prog("Aggregating A_domestic to 17×17…", 48)
    agg     = aggregate_matrix(
        A_data["A"], use_data["x"],
        use_data["va_total"], use_data["va_comp"],
        ind_sector,
    )
    A_nat17 = agg["A_agg"]

    _prog("Building PCE shares…", 54)
    pce_shares = build_pce_shares(use_data["pce"], com_sector,
                                   use_york_weights=True)

    # ── County data ───────────────────────────────────────────────────────────
    _prog("Loading QCEW county data…", 59)
    qcew = load_qcew(qcew_file)

    # ── National employment (BLS CES) ─────────────────────────────────────────
    _prog("Fetching BLS national employment…", 65)
    bls_raw_live = fetch_bls(bls_api_key)
    bls_src      = "live" if bls_raw_live is not None else "fallback_2022"

    bls_raw: dict[str, float] = {}
    if bls_raw_live:
        bls_raw.update(bls_raw_live)
    # Fill any missing series from the 2022 fallback constants
    missing_series = []
    for sid in ALL_SERIES_IDS:
        if sid not in bls_raw:
            bls_raw[sid] = BLS_FALLBACK_2022.get(sid, 0.0)
            missing_series.append(sid)
    if missing_series and bls_raw_live is not None:
        # Live call succeeded but some series were absent in the response
        bls_src = f"live_partial ({len(missing_series)} series from fallback)"

    national_emp = series_to_emp(bls_raw)

    # ── BEA GDP-by-Industry ───────────────────────────────────────────────────
    _prog("Fetching BEA value added and gross output…", 73)
    bea_va_raw  = fetch_bea_table(bea_api_key, "VA")
    bea_go_raw  = fetch_bea_table(bea_api_key, "GO")

    bea_va, bea_va_src, bea_va_diag = _use_or_fallback(
        bea_va_raw, BEA_VA_FALLBACK,
        "bea_value_added.csv", min_plausible_sum=5_000_000,
    )
    bea_output, bea_output_src, bea_output_diag = _use_or_fallback(
        bea_go_raw, BEA_OUTPUT_FALLBACK,
        "bea_output.csv", min_plausible_sum=10_000_000,
    )

    # Compensation derived from VA using stable sector-level comp/VA ratios.
    # BEA removed the standalone compensation table from GDPbyIndustry in 2023.
    bea_comp      = bea_va * _COMP_VA_RATIOS
    bea_comp_src  = bea_va_src + "_derived"
    bea_comp_diag = bea_va_diag

    # ── MPC ───────────────────────────────────────────────────────────────────
    _prog("Fetching BEA NIPA MPC…", 79)
    nat_mpc = fetch_bea_nipa_mpc(bea_api_key) or mpc

    # ── Employment coefficients ───────────────────────────────────────────────
    _prog("Building employment coefficients…", 84)
    emp_coeffs = build_employment_coefficients(
        national_emp, bea_output, bea_comp, bea_va,
        qcew, agg["va_share"],
        mpc=nat_mpc, regional_retention=regional_retention,
    )

    # ── SDP regionalization ───────────────────────────────────────────────────
    _prog("Applying SDP regionalization…", 88)
    reg = regionalize(
        A_nat17, qcew,
        national_emp_total=national_emp.sum(),
        sdp_alpha=sdp_alpha,
        flq_delta=flq_delta,
    )

    # ── Leontief inverses ─────────────────────────────────────────────────────
    _prog("Pre-computing Leontief inverses…", 92)
    inverses = build_leontief_inverses(
        reg["A_york"], emp_coeffs["li_share"],
        pce_shares, emp_coeffs["hh_share"],
    )

    # ── Spending profiles and validation ─────────────────────────────────────
    _prog("Pre-computing spending profiles and validating coefficients…", 96)
    all_profiles = build_all_profiles(A_nat17)
    validation   = validate_coefficients(emp_coeffs)

    _prog("Model ready.", 100)

    import_removed_mean = float((B_total.sum(0) - B_dom.sum(0)).mean())

    # ── Sector-level RPC summary (for diagnostics tab) ─────────────────────
    rpc_summary = [
        {
            "sector":    SECTOR_LABELS[s],
            "lq":        round(float(reg["lq_york"][s]), 4),
            "rpc_sdp":   round(float(reg["rpc_sdp"][s]), 4),
            "rpc_slq":   round(float(reg["rpc_slq"][s]), 4),
            "rpc_flq":   round(float(reg["rpc_flq"][s]), 4),
            "emp_county":round(float(qcew["emp"][s]), 0),
        }
        for s in range(N)
    ]

    return {
        # Matrices
        "A_nat17":    A_nat17,
        "A_york":     reg["A_york"],
        "A_nat15":    A_nat17,        # backwards-compat alias
        # Regionalization
        "lq_york":    reg["lq_york"],
        "rpc_york":   reg["rpc_york"],
        "rpc_sdp":    reg["rpc_sdp"],
        "rpc_slq":    reg["rpc_slq"],
        "rpc_flq":    reg["rpc_flq"],
        "flq_lambda": reg["flq_lambda"],
        "sdp_alpha":  sdp_alpha,
        "flq_delta":  flq_delta,
        "rpc_summary":rpc_summary,    # new: per-sector RPC table
        # Model components
        "pce_shares":  pce_shares,
        "inverses":    inverses,
        "emp_coeffs":  emp_coeffs,
        # National data
        "national_emp": national_emp,
        "bea_output":   bea_output,
        "bea_comp":     bea_comp,
        "bea_va":       bea_va,
        # County data
        "qcew":         qcew,
        # Proprietor employment breakdown (v7.3)
        "emp_ws":       qcew["emp_ws"],    # wage-and-salary only
        "emp_prop":     qcew["emp_prop"],  # proprietors added
        # Pre-computed lookups
        "all_profiles": all_profiles,
        "validation":   validation,
        # Parameters used
        "nat_mpc":           nat_mpc,
        "regional_retention":regional_retention,
        # Data source provenance
        "bea_output_src":  bea_output_src,
        "bea_output_diag": bea_output_diag,
        "bea_comp_src":    bea_comp_src,
        "bea_comp_diag":   bea_comp_diag,
        "bea_va_src":      bea_va_src,
        "bea_va_diag":     bea_va_diag,
        "bls_src":         bls_src,           # new: BLS live/fallback status
        # Diagnostics
        "import_removed_mean":      import_removed_mean,
        "import_leakage_rate":      float(1.0 - reg["rpc_sdp"].mean()),
        "import_leakage_rate_slq":  float(1.0 - reg["rpc_slq"].mean()),
        "import_leakage_rate_flq":  float(1.0 - reg["rpc_flq"].mean()),
        "x_sum":    float(use_data["x"].sum()),
        "ind_sector": ind_sector,
        "com_sector": com_sector,
        "stability":  inverses["stability"],
        # Mutable state
        "scenarios": [],
    }


# ---------------------------------------------------------------------------
# Internal validation helpers
# (Not part of the public API; useful for unit testing)
# ---------------------------------------------------------------------------

def _assert_sector_array(arr: np.ndarray, name: str) -> None:
    """Assert that arr is a 1-D float array of length N."""
    assert isinstance(arr, np.ndarray), f"{name} must be ndarray"
    assert arr.shape == (N,), f"{name} must have shape ({N},); got {arr.shape}"
    assert arr.dtype.kind == "f", f"{name} must be float dtype"


def _assert_square_matrix(mat: np.ndarray, n: int, name: str) -> None:
    """Assert that mat is an (n, n) float array."""
    assert isinstance(mat, np.ndarray), f"{name} must be ndarray"
    assert mat.shape == (n, n), f"{name} must be ({n},{n}); got {mat.shape}"


def _spot_check_sdp(lq_val: float, alpha: float, expected: float, tol: float = 1e-4) -> bool:
    """Return True if SDP formula gives expected value within tolerance."""
    lq = float(lq_val)
    if lq <= 0:
        result = 0.0
    else:
        result = min(lq / (lq + alpha), 1.0)
    return abs(result - expected) < tol


def _spot_check_leontief_identity(A: np.ndarray, tol: float = 1e-10) -> bool:
    """Return True if (I - A) @ inv(I - A) ≈ I within tolerance."""
    n = len(A)
    L = np.linalg.inv(np.eye(n) - A)
    residual = np.max(np.abs((np.eye(n) - A) @ L - np.eye(n)))
    return bool(residual < tol)


# Example unit test sketch (not executed at import time):
#
# def test_sdp_formula():
#     assert _spot_check_sdp(0.0,  0.20, 0.000)
#     assert _spot_check_sdp(1.0,  0.20, 0.833)
#     assert _spot_check_sdp(2.75, 0.20, 0.932)
#
# def test_leontief_identity():
#     A = np.eye(17) * 0.25
#     assert _spot_check_leontief_identity(A)
#
# def test_naics_to_sector():
#     assert naics_to_sector("23")  == 2   # Construction
#     assert naics_to_sector("921") == 15  # Federal government
#     assert naics_to_sector("81")  == 14  # Private other services
